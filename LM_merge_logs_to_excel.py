import os
import re
import sys
import argparse
import pandas as pd
from datetime import datetime
import chardet
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import filedialog
    TK_AVAILABLE = True
except ImportError:
    TK_AVAILABLE = False

# ---------- 工具函数 ----------
def clean_cell(value):
    if value is None:
        return ''
    s = str(value)
    import unicodedata
    keep = {'\t', '\n', '\r'}
    return ''.join(ch for ch in s if not unicodedata.category(ch).startswith('C') or ch in keep)

def split_line(line):
    fields = re.split(r'\s+', line.strip())
    return [f for f in fields if f]

def detect_encoding(filepath):
    with open(filepath, 'rb') as f:
        raw = f.read(10000)
        result = chardet.detect(raw)
        return result['encoding']

def parse_datetime_from_fields(fields):
    if len(fields) < 2:
        return None
    date_str = fields[0].strip()
    time_str = fields[1].strip()
    try:
        dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S.%f")
        return dt
    except ValueError:
        try:
            dt = datetime.strptime(f"{date_str} {time_str}", "%Y/%m/%d %H:%M:%S.%f")
            return dt
        except ValueError:
            return None

def extract_status(fields):
    if len(fields) < 3:
        return ''
    event_text = ' '.join(fields[2:])
    event_lower = event_text.lower()
    if 'down' in event_lower:
        return 'DOWN'
    elif 'run' in event_lower:
        return 'RUN'
    elif 'idle' in event_lower:
        return 'IDLE'
    elif '复位' in event_text:
        return '复位'
    return ''

def extract_reason_id(fields):
    if len(fields) < 3:
        return ''
    event_text = ' '.join(fields[2:])
    match = re.search(r'ReasonID[:=](\d+)', event_text, re.IGNORECASE)
    return match.group(1) if match else ''

def merge_logs_to_excel(folder_path, output_excel, file_extensions=('.log', '.txt')):
    if not os.path.isdir(folder_path):
        print(f"错误：文件夹 '{folder_path}' 不存在")
        return

    all_rows = []
    matched_flags = []
    max_cols = 0
    matched_events = []

    fallback_encodings = ['gbk', 'utf-8', 'big5', 'latin-1']
    keyword_pattern = re.compile(r'(DOWN|复位|RUN|IDLE)', re.IGNORECASE)

    for filename in os.listdir(folder_path):
        if filename.startswith('~$'):
            print(f"跳过临时文件: {filename}")
            continue
        if not filename.lower().endswith(file_extensions):
            continue

        filepath = os.path.join(folder_path, filename)
        try:
            detected_enc = detect_encoding(filepath)
            print(f"文件 {filename} 检测编码: {detected_enc}")
        except Exception:
            detected_enc = None

        encodings_to_try = [detected_enc] + fallback_encodings if detected_enc else fallback_encodings
        lines = None
        for enc in encodings_to_try:
            if enc is None:
                continue
            try:
                with open(filepath, 'r', encoding=enc, errors='strict') as f:
                    lines = f.readlines()
                print(f"  -> 使用编码 {enc} 成功读取")
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if lines is None:
            print(f"跳过文件 {filename}（所有编码尝试均失败）")
            continue

        for line in lines:
            if not line.strip():
                continue
            fields = split_line(line)
            if not fields:
                continue
            cleaned = [clean_cell(f) for f in fields]
            all_rows.append(cleaned)
            max_cols = max(max_cols, len(cleaned))

            is_matched = bool(keyword_pattern.search(line))
            matched_flags.append(is_matched)

            if is_matched:
                dt = parse_datetime_from_fields(cleaned)
                if dt is None:
                    print(f"警告：无法解析日期时间，行内容：{line[:80]}...")
                matched_events.append((dt, cleaned))

    if not all_rows:
        print("未找到任何有效的日志数据。")
        return

    for row in all_rows:
        while len(row) < max_cols:
            row.append('')
    columns = [f'COL_{i+1}' for i in range(max_cols)]
    df_main = pd.DataFrame(all_rows, columns=columns)
    df_main = df_main.astype(str)

    # ---------- 构建 Analysis ----------
    matched_events.sort(key=lambda x: (x[0] is None, x[0]) if x[0] else (True, None))

    temp = []
    for dt, fields in matched_events:
        status = extract_status(fields)
        reason_id = extract_reason_id(fields) if status == 'DOWN' else ''
        temp.append([dt, fields, status, reason_id])

    analysis_cols = list(columns) + ['Status', 'ReasonID', 'Duration (s)']
    analysis_rows = []

    for i, (dt, fields, status, reason_id) in enumerate(temp):
        row_data = fields.copy()
        row_data.append(status)
        row_data.append(reason_id)
        duration = None
        if dt is not None and i + 1 < len(temp):
            next_dt = temp[i+1][0]
            if next_dt is not None:
                duration = (next_dt - dt).total_seconds()
        row_data.append(duration)
        analysis_rows.append(row_data)

    df_analysis = pd.DataFrame(analysis_rows, columns=analysis_cols)

    # ---------- 写入 Excel ----------
    try:
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df_main.to_excel(writer, sheet_name='Merged Data', index=False)
            df_analysis.to_excel(writer, sheet_name='Analysis', index=False)

            workbook = writer.book

            # ---- 标黄 Merged Data ----
            sheet_main = workbook['Merged Data']
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row_idx, is_match in enumerate(matched_flags, start=2):
                if is_match:
                    for col_idx in range(1, len(columns) + 1):
                        cell = sheet_main.cell(row=row_idx, column=col_idx)
                        cell.fill = yellow_fill

            # ---- 在 Analysis 的 N-Q 列添加统计表 ----
            sheet_analysis = workbook['Analysis']

            # 提取统计所需数据
            stats_df = df_analysis[['Status', 'ReasonID', 'Duration (s)']].copy()
            stats_df = stats_df[stats_df['Status'] != '']  # 只统计有状态的行
            stats_df['Duration (s)'] = stats_df['Duration (s)'].fillna(0)

            # 分组统计
            grouped = stats_df.groupby(['Status', 'ReasonID'], as_index=False).agg(
                Times=('Status', 'size'),
                Time_count=('Duration (s)', 'sum')
            )
            grouped.columns = ['Status', 'ReasonID', 'Times', 'Time count']  # 注意列名带空格

            # 自定义排序：RUN, IDLE, DOWN, 复位
            status_order = {'RUN': 1, 'IDLE': 2, 'DOWN': 3, '复位': 4}
            grouped['order'] = grouped['Status'].map(status_order).fillna(5)
            grouped = grouped.sort_values(['order', 'ReasonID']).drop('order', axis=1)

            # 添加总计行
            total_row = pd.DataFrame([['Total', '', grouped['Times'].sum(), grouped['Time count'].sum()]],
                                     columns=grouped.columns)
            grouped = pd.concat([grouped, total_row], ignore_index=True)

            # 写入到 N-Q 列（列索引 14~17）
            start_col = 14  # N列
            start_row = 1   # 第1行

            # 表头
            headers = ['Status', 'ReasonID', 'Times', 'Time count']
            for col_idx, header in enumerate(headers, start=start_col):
                cell = sheet_analysis.cell(row=start_row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # 数据行
            for r_idx, row in enumerate(grouped.itertuples(index=False), start=start_row + 1):
                for c_idx, value in enumerate(row, start=start_col):
                    cell = sheet_analysis.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    if isinstance(value, float):
                        cell.number_format = '0.00'

            # 列宽
            for col_idx in range(start_col, start_col + 4):
                sheet_analysis.column_dimensions[get_column_letter(col_idx)].width = 15

        print(f"✅ 成功生成 Excel 文件：{output_excel}")
        print(f"   - 合并数据行数：{len(df_main)}")
        print(f"   - 匹配事件数：{len(matched_events)}")
        print(f"   - 统计表已放置在 N-Q 列")

    except Exception as e:
        print(f"❌ 写入 Excel 失败: {e}")
        import traceback
        traceback.print_exc()

# ---------- 图形界面 ----------
def gui_select():
    if not TK_AVAILABLE:
        print("错误：tkinter 不可用，请使用命令行模式。")
        sys.exit(1)
    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title="选择存放日志的文件夹")
    if not folder:
        print("未选择输入文件夹。")
        sys.exit()
    output = filedialog.asksaveasfilename(
        title="保存 Excel 文件为",
        defaultextension=".xlsx",
        filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        initialfile="merged_logs.xlsx"
    )
    if not output:
        print("未选择输出文件。")
        sys.exit()
    return folder, output

if __name__ == '__main__':
    os.environ['TK_SILENCE_DEPRECATION'] = '1'
    if len(sys.argv) == 1:
        folder, output = gui_select()
        merge_logs_to_excel(folder, output, ('.log', '.txt'))
    else:
        parser = argparse.ArgumentParser(description='合并日志并按空格拆分，导出Excel，含Analysis')
        parser.add_argument('folder', help='日志文件夹')
        parser.add_argument('-o', '--output', default='merged_logs.xlsx', help='输出Excel文件')
        parser.add_argument('-e', '--extensions', nargs='+', default=['.log', '.txt'], help='文件后缀')
        args = parser.parse_args()
        extensions = tuple(ext.lower() if ext.startswith('.') else f'.{ext.lower()}' for ext in args.extensions)
        merge_logs_to_excel(args.folder, args.output, extensions)